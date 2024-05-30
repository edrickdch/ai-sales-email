from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from langchain_community.chat_models import ChatOllama
from langchain_core.prompts import PromptTemplate
import json
from openpyxl import Workbook, load_workbook


# Load the sales email leads excel file and select the active worksheet
wb = load_workbook("sales-email-leads.xlsx")
ws = wb.active

# Read the header
header = [cell.value for cell in ws[1]]

# Read the data
leads_data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    leads_data.append(dict(zip(header, row)))

# Set up Chrome options for scraping the /about page
options = Options()
options.add_argument("--headless")  # Run in headless mode (no GUI)
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")

# Path to your ChromeDriver
webdriver_service = Service(ChromeDriverManager().install())

# Initialize WebDriver
driver = webdriver.Chrome(service=webdriver_service, options=options)

# Load the website
url = leads_data[0]["website"]
driver.get(url)

# Wait for JavaScript to render the content
time.sleep(2)  # Adjust sleep time as necessary for the page to load

# Extract the HTML content
html_content = driver.page_source

# Parse the HTML with BeautifulSoup
soup = BeautifulSoup(html_content, "html.parser")

# Extract meaningful text
text_elements = soup.find_all(string=True)

# Filter text elements to remove non-meaningful content
meaningful_text = []
for element in text_elements:
    if element.parent.name not in [
        "script",
        "style",
        "head",
        "title",
        "meta",
        "[document]",
    ]:
        content = element.strip()
        if content:
            meaningful_text.append(content)

# Join the meaningful text into a single string
cleaned_text = " ".join(meaningful_text)

# Close the WebDriver
driver.quit()


# Use llama3
llm = ChatOllama(model="llama3", temperature=0)

# Prompt to summarize the about page
summarize_prompt = PromptTemplate(
    template="""
    <|begin_of_text|><|start_header_id|>system<|end_header_id|>
    You are a summarizing assistant. Your task is to summarize a company's about page.
    The data has been cleaned of html, css, js and other such kind of tags.
    However there is still noise in the scraped about page that relates to these.
    Ignore them, and provide a summary of relevant information of the company specifically.  
    <|eot_id|><|start_header_id|>user<|end_header_id|>
    Here is the parsed about page of the company: \n\n {document} \n
    <|eot_id|><|start_header_id|>assistant<|end_header_id|>
    """.strip(),
    input_variables=["document"],
)

# Summarize chain
summarizer = summarize_prompt | llm

# Stream /about summary generation
# For demonstration purpose, we are printing to console token by token and flushing write buffer frequently
summarized_output_chunks = []
with open("about-summarized-output.md", "w", buffering=1) as f:
    for data in summarizer.stream({"document": cleaned_text}):
        print(data.content)
        f.write(data.content)
        summarized_output_chunks.append(data.content)

# Recover the output
summarized_output = "".join(summarized_output_chunks)

# Prompt to craft the email based on the /about page and the email intent
email_crafting_prompt = PromptTemplate(
    template="""
    <|begin_of_text|><|start_header_id|>system<|end_header_id|>
    You are a skilled saleswoman and email assistant. Your task is to create an outreach email. 
    You are given a summary of the company's about page.
    Keep it short and simple. Use very basic words and sentences.
    Write a very short intro email, show that you've researched them and ask if they have 15 minutes to chat in the coming weeks.
    Add that you are looking forward to hear from them and work together.
    Append you're signature at the end, as John Doe
    ONLY OUTPUT THE EMAIL MESSAGE
    DO NOT PRESENT THE EMAIL
    DO NOT WRITE A SUBJECT LINE
    <|eot_id|><|start_header_id|>user<|end_header_id|>
    Here is company's summarized about page: \n\n {document} \n
    Here is our company's objective for the outreach: \n\n {objective} \n
    Here is the name of the person you are reaching out to: \n\n {name} \n
    <|eot_id|><|start_header_id|>assistant<|end_header_id|>
    """.strip(),
    input_variables=["document", "objective", "name"],
)

# Email creation chain
email_crafter = email_crafting_prompt | llm

# Stream summary
# For demonstration purpose, we are printing to console token by token and flushing write buffer frequently
email_output_chunks = []
with open("email-output.md", "w", buffering=1) as f:
    for data in email_crafter.stream(
        {
            "document": summarized_output,
            "objective": leads_data[0]["objective"],
            "name": leads_data[0]["name"],
        }
    ):
        print(data.content)
        f.write(data.content)
        email_output_chunks.append(data.content)

# Recover the output
email_message = "".join(email_output_chunks)

# Prompt to generate the subject line
subject_line_prompt = PromptTemplate(
    template="""
    <|begin_of_text|><|start_header_id|>system<|end_header_id|>
    You are a skilled email assistant. Your task is to create a subject line. 
    You like to keep things short and simple.
    You are only allowed to use 6 words maximum.
    ONLY OUTPUT THE SUBJECT LINE
    <|eot_id|><|start_header_id|>user<|end_header_id|>
    Here is the email: \n\n {document} \n
    <|eot_id|><|start_header_id|>assistant<|end_header_id|>
    """.strip(),
    input_variables=["document"],
)

# Subject line creation chain
subject_line_crafter = subject_line_prompt | llm

# Run the creation chain to generate the subject line
subject_line = subject_line_crafter.invoke({"document": email_message}).content

# Format email data
email_data = {
    "name": leads_data[0]["name"],
    "email": leads_data[0]["email"],
    "subject_line": subject_line,
    "message": email_message,
}

# Save the output to JSON
with open("final-output.json", "w") as f:
    json.dump(email_data, f, indent=4)


# Save the output to excel format
wb = Workbook()
ws = wb.active

# Write the header
ws.append(list(email_data.keys()))

# Write the data
ws.append(list(email_data.values()))

# Save the workbook to a file
wb.save("output.xlsx")
